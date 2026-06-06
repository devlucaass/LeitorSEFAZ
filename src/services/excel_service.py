from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from tkinter import messagebox
from config.config_manager import load_config

config = load_config()


class ExcelBot:

    def __init__(self, dados):
        self.dados = dados

    def inserir_no_excel(self, caminho_excel, celula_inicial):

        try:
            wb = load_workbook(caminho_excel)
            nome_aba = config['nome_planilha']

            if nome_aba not in wb.sheetnames:
                messagebox.showerror('Erro', 'Planilha não encontrada.')
                return

            ws = wb[nome_aba]
            coluna_letra, linha_excel = coordinate_from_string(celula_inicial)
            coluna_inicial = column_index_from_string(coluna_letra)

            for linha in self.dados:
                coluna_excel = coluna_inicial

                valores = [
                    linha['data'],
                    linha['estabelecimento'],
                    linha['produto'],
                    linha['grandeza'],
                    linha['quantidade_comprada'],
                    linha['valor_unitario'],
                    linha['quantidade'],
                    linha['valor_total']
                ]

                for valor in valores:
                    celula = ws.cell(row=linha_excel, column=coluna_excel, value=valor)
                    celula._style = copy(ws.cell(row=2,column=coluna_excel)._style)
                    celula.alignment = Alignment(horizontal='center', vertical='center')

                    if coluna_excel in [6, 8]:
                        celula.number_format = 'R$ #,##0.00'

                    elif coluna_excel == 7:
                        celula.number_format = '0.000'

                    coluna_excel += 1

                linha_excel += 1

            wb.save(caminho_excel)

            messagebox.showinfo('Sucesso', 'Dados inseridos com sucesso!')

        except Exception as erro:
            messagebox.showerror('Erro', f'Ocorreu um erro:\n{erro}')