import os

import xlwings as xw
from openpyxl import Workbook, load_workbook

from config.config_manager import ConfigManager
from config.constants import EXCEL_PATH_KEY, SHEET_NAME_KEY
from utils.file_dialog import save_file


class ExcelBot:
    def __init__(self, data):
        self.data = data

    def insert_data(self, excel_cell):
        try:
            config = ConfigManager.load_config()

            excel_path = config.get(EXCEL_PATH_KEY)
            sheet_name = config.get(SHEET_NAME_KEY)

            if not excel_path or not sheet_name:
                return False

            wb = self._open_workbook(excel_path)
            ws = self._get_worksheet(wb, sheet_name)

            excel_values = self._prepare_excel_values()

            ws.range(excel_cell).value = excel_values

            wb.save()

            return True

        except (FileNotFoundError, KeyError, PermissionError, OSError):
            return False

    @staticmethod
    def create_spreadsheet_template():
        try:
            config = ConfigManager.load_config()

            excel_path = save_file()

            if not excel_path:
                return False

            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active

            else:
                wb = Workbook()
                ws = wb.active
                ws.title = config[SHEET_NAME_KEY]

            ExcelBot._create_columns(ws)
            wb.save(excel_path)

            return True

        except (PermissionError, KeyError, OSError):
            return False

    @staticmethod
    def _create_columns(ws):
        columns = [
            "DATA",
            "ESTABELECIMENTO",
            "PRODUTO",
            "GRANDEZA",
            "QUANTIDADE P/ PRODUTO",
            "VALOR UNITÁRIO",
            "QUANTIDADE",
            "TOTAL",
        ]

        for column, value in enumerate(columns, start=1):
            ws.cell(row=1, column=column, value=value)

    def _open_workbook(self, excel_path):
        return xw.Book(excel_path)

    def _get_worksheet(self, wb, sheet_name):
        return wb.sheets[sheet_name]

    def _prepare_excel_values(self):
        return [
            [
                row["date"],
                row["establishment_name"],
                row["product"],
                row["unit"],
                row["purchased_quantity"],
                row["unit_price"],
                row["quantity"],
                row["total_price"],
            ]
            for row in self.data
        ]
